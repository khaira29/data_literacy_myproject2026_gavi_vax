import pandas as pd
from openpyxl import load_workbook
import pdfplumber
import pandas as pd

PDF_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/00_raw_data/gavi_eligibility_country.pdf"
OUT_EXCEL = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/01_interm_data/gavi_eligibility_country.xlsx"

rows = []

with pdfplumber.open(PDF_FILE) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                rows.append(row)

df = pd.DataFrame(rows, columns=["country", "year", "gavi_eligibility_group"])

# Clean
df = df.dropna(subset=["country", "year"])
df["year"] = pd.to_numeric(df["year"], errors="coerce")

df.to_excel(OUT_EXCEL, index=False)

print("Saved:", OUT_EXCEL)
print(df.head())

# --------------------------------------------------
# INPUT / OUTPUT
# --------------------------------------------------
INPUT_1 = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/00_raw_data/gavi_eligibility_country.xlsx"
SHEET_1 = "Table 1"

INPUT_2 = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/00_raw_data/gavi_mic_countries.xlsx"

OUTPUT_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/01_interm_data/gavi_eligibility_country_wide.xlsx"

MIC_FILL_START_YEAR = 2022
MIC_FILL_END_YEAR   = 2025

# --------------------------------------------------
# Step 1: Build WIDE table from INPUT_1
# --------------------------------------------------
df = pd.read_excel(INPUT_1, sheet_name=SHEET_1, engine="openpyxl")

df = df[["Country", "Year", "Gavi eligibility group"]].rename(columns={
    "Country": "country_name",
    "Year": "year",
    "Gavi eligibility group": "gavi_eligibility_group",
})

df["country_name"] = df["country_name"].astype(str).str.strip()
df["year"] = pd.to_numeric(df["year"], errors="coerce")
df["gavi_eligibility_group"] = df["gavi_eligibility_group"].astype(str).str.strip()

df = df.dropna(subset=["country_name", "year"])
df["year"] = df["year"].astype(int)

# Keep FIRST occurrence if duplicates exist
df = df.drop_duplicates(subset=["country_name", "year"], keep="first")

wide = df.pivot(index="country_name", columns="year", values="gavi_eligibility_group")
wide.columns = [f"gavi_{y}" for y in wide.columns]
wide = wide.reset_index()

# --------------------------------------------------
# Step 2: Load MIC input (INPUT_2)
# --------------------------------------------------
mic = pd.read_excel(INPUT_2, engine="openpyxl")
mic = mic[["country_name", "gavi_mic_status"]].copy()
mic["country_name"] = mic["country_name"].astype(str).str.strip()
mic["gavi_mic_status"] = mic["gavi_mic_status"].astype(str).str.strip()

# --------------------------------------------------
# Step 3: Ensure wide has columns gavi_2022 ... gavi_2025
# --------------------------------------------------
for y in range(MIC_FILL_START_YEAR, MIC_FILL_END_YEAR + 1):
    col = f"gavi_{y}"
    if col not in wide.columns:
        wide[col] = pd.NA

# --------------------------------------------------
# Step 4: PRINT CHECKS FIRST (before overwriting)
# --------------------------------------------------
wide_names = set(wide["country_name"])
mic_names = set(mic["country_name"])

new_in_mic = sorted(mic_names - wide_names)
existing_in_mic = sorted(mic_names & wide_names)

print("\n=== MIC country name check ===")
print(f"MIC countries total: {len(mic_names)}")
print(f"Already in wide data: {len(existing_in_mic)}")
print(f"NEW (not in wide data): {len(new_in_mic)}")

print("\nCountries already in wide data (and also in MIC list):")
for c in existing_in_mic:
    print(c)

years_to_show = [2022, 2023, 2024, 2025]
cols_to_show = ["country_name"] + [f"gavi_{y}" for y in years_to_show]

print("\n=== CURRENT wide values BEFORE overwrite (2022–2025) for existing MIC countries ===")
df_existing_mic_before = (
    wide.loc[wide["country_name"].isin(existing_in_mic), cols_to_show]
    .sort_values("country_name")
)
print(df_existing_mic_before.to_string(index=False))

# --------------------------------------------------
# EXTRA CHECK: counts of MIC countries
# (BEFORE overwriting any values)
# --------------------------------------------------
wide_names = set(wide["country_name"])
mic_names = set(mic["country_name"])

existing_in_wide = sorted(mic_names & wide_names)
new_not_in_wide  = sorted(mic_names - wide_names)

print("\n=== PRE-OVERWRITE MIC COVERAGE CHECK ===")
print(f"Total MIC countries           : {len(mic_names)}")
print(f"MIC countries already in wide : {len(existing_in_wide)}")
print(f"MIC countries NOT in wide     : {len(new_not_in_wide)}")

print("\nMIC countries already in wide data:")
for c in existing_in_wide:
    print(c)

print("\nMIC countries NOT in wide data (will be added):")
for c in new_not_in_wide:
    print(c)


# --------------------------------------------------
# Step 5: Ensure ALL MIC countries exist in wide (append missing ones)
# --------------------------------------------------
if new_in_mic:
    add_rows = pd.DataFrame({"country_name": new_in_mic})
    # create empty year cols (only 2022–2025 are needed for MIC overwrite)
    for y in range(MIC_FILL_START_YEAR, MIC_FILL_END_YEAR + 1):
        add_rows[f"gavi_{y}"] = pd.NA
    wide = pd.concat([wide, add_rows], ignore_index=True)

# --------------------------------------------------
# Step 6: Overwrite gavi_2022–gavi_2025 with MIC status for ALL MIC countries
# --------------------------------------------------
mic_map = dict(zip(mic["country_name"], mic["gavi_mic_status"]))

mic_val = wide["country_name"].map(mic_map)

for y in range(MIC_FILL_START_YEAR, MIC_FILL_END_YEAR + 1):
    col = f"gavi_{y}"
    mask = mic_val.notna()  # only MIC countries
    wide.loc[mask, col] = mic_val[mask].values

# --------------------------------------------------
# Final formatting: order columns
# --------------------------------------------------
year_cols_sorted = sorted(
    [c for c in wide.columns if c.startswith("gavi_")],
    key=lambda x: int(x.split("_")[1])
)

wide = wide[["country_name"] + year_cols_sorted].copy()
wide = wide.sort_values("country_name").reset_index(drop=True)

# --------------------------------------------------
# Save
# --------------------------------------------------
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    wide.to_excel(writer, sheet_name="gavi_wide_plus_mic", index=False)

print("\nSaved:", OUTPUT_FILE)
print("Countries:", wide["country_name"].nunique())
print("Year columns:", len(year_cols_sorted))
print("First–last year:", year_cols_sorted[0], year_cols_sorted[-1])

# --------------------------------------------------
# Optional: Verify after overwrite
# --------------------------------------------------
print("\n=== wide values AFTER overwrite (2022–2025) for all MIC countries ===")
df_mic_after = (
    wide.loc[wide["country_name"].isin(sorted(mic_names)), ["country_name"] + [f"gavi_{y}" for y in years_to_show]]
    .sort_values("country_name")
)
print(df_mic_after.to_string(index=False))


# --------------------------------------------------
# ADDITIONAL OUTPUT (ENHANCED):
# Create sheet "gavi_country_2024" with country_code + country_name + gavi_2024
# Prefill some country_code, then map remaining via income_class_2024, then hpv_vax_2024
# --------------------------------------------------
FINAL_FILE  = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/02_cleaned_data/dl_project_section_1.xlsx"
FINAL_SHEET = "gavi_country_2024"

SHEET_INC = "income_class_2024"
SHEET_VAX = "hpv_vax_2024"

# 1) Base: from already-built `wide`
df_2024 = wide[["country_name", "gavi_2024"]].copy()
df_2024["country_name"] = df_2024["country_name"].astype("string").str.strip()

# Create empty country_code
df_2024["country_code"] = pd.NA

# 2) Prefill tricky codes (provided list)
prefill_map = {
    "Bolivia, Plurinational State of": "BOL",
    "Congo": "COG",
    "Congo, Democratic Republic of": "COD",
    "Egypt": "EGY",
    "Iran": "IRN",
    "Korea DPR": "PRK",
    "Micronesia": "FSM",
    "Moldova, Republic of": "MDA",
    "Occupied Palestinian territory": "PSE",  # ISO alpha-3 = PSE :contentReference[oaicite:1]{index=1}
    "Somalia": "SOM",
    "Tanzania, United Republic of": "TZA",
    "Venezuela": "VEN",
    "Yemen": "YEM",
}

df_2024.loc[df_2024["country_name"].isin(prefill_map.keys()), "country_code"] = (
    df_2024["country_name"].map(prefill_map)
)

# 3) Load reference sheets for mapping
inc = pd.read_excel(FINAL_FILE, sheet_name=SHEET_INC, engine="openpyxl")
vax = pd.read_excel(FINAL_FILE, sheet_name=SHEET_VAX, engine="openpyxl")

def make_name_to_code(df, label):
    tmp = df[["country_code", "country_name"]].copy()
    tmp["country_code"] = tmp["country_code"].astype("string").str.strip().str.upper()
    tmp["country_name"] = tmp["country_name"].astype("string").str.strip()

    # keep first code per name (name→code may not be unique; follow "keep first" convention)
    tmp = tmp.dropna(subset=["country_name", "country_code"])
    tmp = tmp.drop_duplicates(subset=["country_name"], keep="first")

    print(f"\n=== Mapping table size: {label} ===")
    print("Unique country_name:", tmp["country_name"].nunique())
    print("Unique country_code:", tmp["country_code"].nunique())

    return dict(zip(tmp["country_name"], tmp["country_code"]))

inc_map = make_name_to_code(inc, "income_class_2024")
vax_map = make_name_to_code(vax, "hpv_vax_2024")

# 4) Map remaining: INCOME first
mask_need = df_2024["country_code"].isna()
df_2024.loc[mask_need, "country_code"] = df_2024.loc[mask_need, "country_name"].map(inc_map)

mapped_by_income = int(df_2024["country_code"].notna().sum())

# 5) Map still-unmatched: VAX second
mask_need = df_2024["country_code"].isna()
df_2024.loc[mask_need, "country_code"] = df_2024.loc[mask_need, "country_name"].map(vax_map)

# 6) Print diagnostics BEFORE writing anything
total = len(df_2024)
mapped_total = int(df_2024["country_code"].notna().sum())
unmapped = total - mapped_total

print("\n=== gavi_country_2024 country_code mapping diagnostics ===")
print("Total rows:", total)
print("Mapped after prefill:", int(df_2024["country_name"].isin(prefill_map).sum()))
print("Mapped after income step (cumulative):", mapped_by_income)
print("Mapped after vax step (cumulative):", mapped_total)
print("Still unmapped:", unmapped)

if unmapped > 0:
    print("\n--- Still-unmapped country_name values ---")
    print(
        df_2024.loc[df_2024["country_code"].isna(), ["country_name", "gavi_2024"]]
        .sort_values("country_name")
        .to_string(index=False)
    )

# Optional: enforce format
df_2024["country_code"] = df_2024["country_code"].astype("string").str.upper().str.strip()

# Reorder columns nicely
df_2024 = df_2024[["country_code", "country_name", "gavi_2024"]].sort_values(
    ["country_code", "country_name"], na_position="last"
).reset_index(drop=True)

# 7) Write the sheet (replace if exists)
wb = load_workbook(FINAL_FILE)
if FINAL_SHEET in wb.sheetnames:
    wb.remove(wb[FINAL_SHEET])
    wb.save(FINAL_FILE)
wb.close()

with pd.ExcelWriter(FINAL_FILE, engine="openpyxl", mode="a") as writer:
    df_2024.to_excel(writer, sheet_name=FINAL_SHEET, index=False)

print("\nSaved 2024 Gavi status (with country_code) to:")
print(FINAL_FILE)
print("Sheet name:", FINAL_SHEET)
print("Rows:", len(df_2024))

