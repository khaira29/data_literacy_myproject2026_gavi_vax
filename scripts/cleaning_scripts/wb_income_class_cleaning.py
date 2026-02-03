"""
Combine CODE 1 + CODE 2 into ONE script:

1) Read WB historical income workbook (raw) and write a cleaned file with 2015–2024.
2) From that cleaned file, keep (country_code, country_name, 2024), rename 2024 -> income_class.
3) Append/replace a sheet "income_class_2024" into dl_project_section_1.xlsx.

Requirements:
pip install openpyxl pandas
"""

import pandas as pd
from openpyxl import load_workbook, Workbook

# ==================================================
# PATHS
# ==================================================
RAW_INPUT_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/00_raw_data/wb_hist_income_country.xlsx"
INTERM_OUTPUT_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/01_interm_data/wb_hist_income_2015_2024_clean.xlsx"

FINAL_BOOK = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/02_cleaned_data/dl_project_section_1.xlsx"
FINAL_SHEET = "income_class_2024"

# ==================================================
# SETTINGS FOR EXTRACTION (from RAW_INPUT_FILE)
# ==================================================
SHEET_NAME = "Country Analytical History"

INPUT_START_ROW = 12
COL_COUNTRY_CODE = 1    # A
COL_COUNTRY_NAME = 2    # B
COL_YEAR_START = 31     # AE (2015 in file layout)

START_YEAR = 2015
N_YEARS = 10            # 2015–2024
STOP_CODE = "ZWE"       # include row then stop

# ==================================================
# STEP 1: Create intermediate cleaned workbook (2015–2024)
# ==================================================
wb_in = load_workbook(RAW_INPUT_FILE, data_only=True)
if SHEET_NAME not in wb_in.sheetnames:
    raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb_in.sheetnames}")
ws_in = wb_in[SHEET_NAME]

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "clean_2015_2024"

# headers
ws_out.cell(row=1, column=1).value = "country_code"
ws_out.cell(row=1, column=2).value = "country_name"
for i in range(N_YEARS):
    ws_out.cell(row=1, column=3 + i).value = START_YEAR + i

out_row = 2

for r in range(INPUT_START_ROW, ws_in.max_row + 1):
    country_code = ws_in.cell(row=r, column=COL_COUNTRY_CODE).value

    if country_code in (None, ""):
        continue

    ws_out.cell(row=out_row, column=1).value = country_code
    ws_out.cell(row=out_row, column=2).value = ws_in.cell(row=r, column=COL_COUNTRY_NAME).value

    for i in range(N_YEARS):
        ws_out.cell(row=out_row, column=3 + i).value = ws_in.cell(
            row=r,
            column=COL_YEAR_START + i
        ).value

    if country_code == STOP_CODE:
        break

    out_row += 1

wb_out.save(INTERM_OUTPUT_FILE)
wb_in.close()

print("Saved intermediate cleaned file:", INTERM_OUTPUT_FILE)

# ==================================================
# STEP 2: Build income_class_2024 dataframe from intermediate file
# ==================================================
df = pd.read_excel(INTERM_OUTPUT_FILE, engine="openpyxl")

year_col = 2024 if 2024 in df.columns else "2024"
if year_col not in df.columns:
    raise ValueError(
        f"Year column 2024 not found. Available columns: {list(df.columns)}"
    )

df_out = df[["country_code", "country_name", year_col]].copy()
df_out = df_out.rename(columns={year_col: "income_class"})

# ==================================================
# STEP 3: Append/replace the sheet in FINAL_BOOK
# ==================================================
wb = load_workbook(FINAL_BOOK)
if FINAL_SHEET in wb.sheetnames:
    wb.remove(wb[FINAL_SHEET])
    wb.save(FINAL_BOOK)
wb.close()

with pd.ExcelWriter(FINAL_BOOK, engine="openpyxl", mode="a") as writer:
    df_out.to_excel(writer, sheet_name=FINAL_SHEET, index=False)

print("Saved sheet to existing workbook:")
print("Workbook:", FINAL_BOOK)
print("Sheet:", FINAL_SHEET)
print("Columns:", df_out.columns.tolist())
print("Rows:", len(df_out))
