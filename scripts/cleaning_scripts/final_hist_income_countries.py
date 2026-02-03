"""
Create intermediate WB historical income file (LONG format, 2008–2024)

- Reads RAW_INPUT_FILE
- Extracts country_code, country_name, income class by year
- Outputs LONG format:
    country_code | country_name | year | income_class
- Keeps income_class only if in {H, L, LM, UM}, else blank
"""

from openpyxl import load_workbook, Workbook

# ==================================================
# PATHS
# ==================================================
RAW_INPUT_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/00_raw_data/wb_hist_income_country.xlsx"
INTERM_OUTPUT_FILE = r"/Users/khaira_abdillah/Documents/dl_pro_country_comp/01_interm_data/final_wb_hist_income.xlsx"

# ==================================================
# SETTINGS
# ==================================================
SHEET_NAME = "Country Analytical History"

INPUT_START_ROW = 12
COL_COUNTRY_CODE = 1    # A
COL_COUNTRY_NAME = 2    # B
COL_YEAR_START = 24     # X (2008)

START_YEAR = 2008
N_YEARS = 17            # 2008–2024 inclusive
STOP_CODE = "ZWE"

VALID_INCOME = {"H", "L", "LM", "UM"}

# ==================================================
# LOAD RAW WORKBOOK
# ==================================================
wb_in = load_workbook(RAW_INPUT_FILE, data_only=True)
if SHEET_NAME not in wb_in.sheetnames:
    raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb_in.sheetnames}")
ws_in = wb_in[SHEET_NAME]

# ==================================================
# CREATE OUTPUT (LONG FORMAT)
# ==================================================
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "long_2008_2024"

# headers
ws_out.append(["country_code", "country_name", "year", "income_class"])

out_row = 2

for r in range(INPUT_START_ROW, ws_in.max_row + 1):
    country_code = ws_in.cell(row=r, column=COL_COUNTRY_CODE).value

    if country_code in (None, ""):
        continue

    country_name = ws_in.cell(row=r, column=COL_COUNTRY_NAME).value

    for i in range(N_YEARS):
        year = START_YEAR + i
        raw_val = ws_in.cell(
            row=r,
            column=COL_YEAR_START + i
        ).value

        # --- sanitize income class ---
        if raw_val is None:
            income_val = None
        else:
            val = str(raw_val).strip().upper()
            income_val = val if val in VALID_INCOME else None

        ws_out.cell(row=out_row, column=1).value = country_code
        ws_out.cell(row=out_row, column=2).value = country_name
        ws_out.cell(row=out_row, column=3).value = year
        ws_out.cell(row=out_row, column=4).value = income_val

        out_row += 1

    if country_code == STOP_CODE:
        break

wb_out.save(INTERM_OUTPUT_FILE)
wb_in.close()

print("Saved LONG-format intermediate file:", INTERM_OUTPUT_FILE)
print("Rows written:", out_row - 2)
print("Years:", START_YEAR, "-", START_YEAR + N_YEARS - 1)
